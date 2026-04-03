"""
engine/report_builder.py  --  Excel Report Generator (Server)
=============================================================
Generates the full 10-sheet Excel report in /tmp (server-safe).
Returns bytes for Streamlit download_button.
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

log = logging.getLogger(__name__)

# Colour palette
DARK   = "1A1A2E"; HEADER = "0F3460"
G      = "1D9E75"; A      = "BA7517"; R = "A32D2D"
BG1    = "F2F4F8"; BG2    = "FFFFFF"


def _fill(h):  return PatternFill("solid", fgColor=h)
def _fnt(bold=False, color=DARK, size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def _aln(h="center"): return Alignment(horizontal=h, vertical="center")
def _bdr():
    s = Side(style="thin", color="CCCCDD")
    return Border(left=s, right=s, top=s, bottom=s)
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h


def wcell(ws, r, c, v, bold=False, color=DARK, bg=None, fmt=None, size=10):
    x = ws.cell(r, c, v)
    x.font = _fnt(bold, color, size); x.alignment = _aln(); x.border = _bdr()
    if bg:  x.fill          = _fill(bg)
    if fmt: x.number_format = fmt
    return x


def titrow(ws, row, text, ncols=9, bg=DARK, h=30):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    x = ws[f"A{row}"]; x.value = text
    x.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    x.fill = _fill(bg); x.alignment = _aln(); rh(ws, row, h)


def hdrrow(ws, row, labels, bg=HEADER, h=22):
    rh(ws, row, h)
    for c, lbl in enumerate(labels, 1):
        x = ws.cell(row, c, lbl)
        x.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        x.fill = _fill(bg); x.alignment = _aln(); x.border = _bdr()


def addimg(ws, path_or_bytes, anchor, w=650, h=380):
    try:
        im = XLImage(path_or_bytes); im.width = w; im.height = h
        ws.add_image(im, anchor)
    except Exception:
        pass

def _add_macro_sheet(wb, macro: dict, regime: dict, gauges: dict):
    """Add 'Macro Pulse' sheet to the Excel workbook."""
    ws = wb.create_sheet("Macro Pulse")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "534AB7"

    for i, w in enumerate([22, 22, 18, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    titrow(ws, 1, f"MACRO PULSE  --  Generated {macro.get('fetched_at','')}", ncols=6)

    # Regime summary
    hdrrow(ws, 2, ["Regime", "Score", "Cash %", "Cash Guidance", "Action", "Generated"])
    rc = {"Defensive": R, "Neutral": A, "Aggressive": G}.get(regime.get("regime",""), DARK)
    for c, v in enumerate([
        regime.get("regime",""), str(regime.get("score","")) + "/10",
        regime.get("cash_pct",""), regime.get("cash_thb_est",""),
        regime.get("action",""), macro.get("fetched_at",""),
    ], 1):
        wcell(ws, 3, c, v, bold=(c == 1), color=("FFFFFF" if c == 1 else DARK),
              bg=(rc if c == 1 else BG1))

    # Key metrics table
    rh(ws, 5, 22)
    hdrrow(ws, 5, ["Indicator", "Value", "Signal", "Note", "", ""])
    rows = [
        ("Thai Policy Rate",    f"{macro['rates']['thai_rate']:.2f}%",    macro['rates']['thai_signal'],     f"as of {macro['rates']['thai_rate_date']}"),
        ("US Fed Rate",         f"{macro['rates']['us_fed_rate']:.2f}%",  macro['rates']['us_signal'],       f"as of {macro['rates']['us_fed_date']}"),
        ("VIX",                 f"{macro['vix'].get('current','n/a'):.1f}" if macro['vix'].get('current') else "n/a", macro['vix'].get('signal','unknown'), "30d change: " + (f"{macro['vix'].get('change_30d',0):+.1f}" if macro['vix'].get('change_30d') else "n/a")),
        ("WTI Oil ($/bbl)",     f"${macro['oil'].get('current','n/a'):.2f}" if macro['oil'].get('current') else "n/a", macro['oil'].get('signal','unknown'), macro['oil'].get('geopolitical_note','')),
        ("USD/THB FX Signal",   f"{macro['fx'].get('current',0):.4f}",    macro['fx'].get('signal','unknown'), macro['fx'].get('advice','')),
        ("Recession Prob.",     f"~{macro['recession'].get('probability','?')}%", macro['recession'].get('signal','unknown'), macro['recession'].get('note','')),
        ("2s10s Spread",        f"{macro['yield_curve'].get('spread_2s10s',0):+.3f}%" if macro['yield_curve'].get('spread_2s10s') is not None else "n/a", macro['yield_curve'].get('signal','unknown'), "Inverted" if macro['yield_curve'].get('inverted') else "Not inverted"),
        ("Credit Risk Score",   f"{gauges.get('default_risk','?')}/100",  "red" if gauges.get('default_risk',0) > 65 else ("yellow" if gauges.get('default_risk',0) > 35 else "green"), "HYG vs LQD proxy"),
    ]
    sig_color = {"green": G, "yellow": A, "red": R, "unknown": "888888"}
    for i, (ind, val, sig, note) in enumerate(rows):
        r = 6 + i; bg = BG1 if i % 2 == 0 else BG2
        wcell(ws, r, 1, ind, bold=True, bg=bg)
        wcell(ws, r, 2, val, bg=bg)
        sc = sig_color.get(sig, "888888")
        xc = wcell(ws, r, 3, sig.upper(), bg=bg, color="FFFFFF")
        xc.fill = _fill(sc); xc.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        wcell(ws, r, 4, note, bg=bg)

def build_report(
    pnl_df:       pd.DataFrame,
    risk_tbl:     pd.DataFrame,
    w_dict:       dict,
    div_rows:     list,
    wht_records:  list,
    bt_metrics:   dict,
    milestones:   dict,
    fx_signal:    dict,
    cal_events:   list,
    cfg:          dict,
    conf:         dict,
    plots:        dict,
    bl_applied:   bool,
) -> bytes:
    """Build full Excel report and return as bytes for download."""

    wb = Workbook(); wb.remove(wb.active)
    meta = cfg.get("meta", {})
    snap = cfg.get("ks_app_snapshot_20260327", {})

    # ── Sheet 1: Dashboard ───────────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = G
    for i, w in enumerate([12, 20, 14, 14, 14, 14, 14, 14, 14], 1): cw(ws, i, w)
    titrow(ws, 1, f"PORTFOLIO ANALYTICS  |  {meta.get('account_holder','')}  "
                  f"|  {meta.get('account_id','')}  |  {datetime.today().strftime('%Y-%m-%d')}")
    kv = [
        ("Market Value (THB)",  f"฿{snap.get('market_value_thb',0):,.2f}"),
        ("Total Cost (THB)",    f"฿{snap.get('total_cost_thb',0):,.2f}"),
        ("Unrealised (THB)",    f"฿{snap.get('unrealized_thb',0):,.2f}"),
        ("Dividends (THB)",     f"฿{snap.get('total_dividends_thb',0):,.2f}"),
        ("Cash (USD)",          f"${conf['cash_usd']:.2f}"),
        ("FX Rate",             f"{conf['fx']:.4f} THB/USD"),
        ("WHT Rate",            f"{conf['wht']*100:.0f}% {'(treaty 15%)' if conf['wht']<0.20 else '(default 30%)'}"),
        ("BL Views Applied",    "YES" if bl_applied else "NO"),
        ("FX Signal",           fx_signal.get("signal", "n/a").replace("_", " ").title()),
        ("Generated",           datetime.today().strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (k, v) in enumerate(kv):
        r = 3 + i; bg = BG1 if i % 2 == 0 else BG2; rh(ws, r, 19)
        wcell(ws, r, 1, k, bold=True, bg=bg)
        ws.merge_cells(f"B{r}:I{r}")
        xc = ws.cell(r, 2, v); xc.fill = _fill(bg); xc.border = _bdr()
        col = G if any(c in v for c in ["฿", "$", "YES"]) else (R if "-" in v else DARK)
        xc.font = _fnt(bold=True, color=col); xc.alignment = _aln()

    # ── Sheet 2: Position P&L ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Position P&L")
    ws2.sheet_view.showGridLines = False; ws2.sheet_properties.tabColor = G
    for i, w in enumerate([6,14,10,14,14,16,16,14,14,12], 1): cw(ws2, i, w)
    titrow(ws2, 1, f"POSITION P&L  --  {datetime.today().strftime('%Y-%m-%d')}")
    hdrrow(ws2, 2, ["#","Ticker","Shares","Avg Cost $","Current $",
                     "Cost Basis $","Market Value $","Unrealised $","Unrealised THB","Net P&L %"])
    for i, (tkr, row) in enumerate(pnl_df.iterrows()):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws2, r, 19)
        pc = G if row["Net P&L $"] >= 0 else R
        wcell(ws2,r,1,i+1,bg=bg); wcell(ws2,r,2,tkr,bold=True,bg=bg,color=HEADER)
        wcell(ws2,r,3,row["Shares"],bg=bg,fmt="#,##0")
        wcell(ws2,r,4,row["Avg Cost $"],bg=bg,fmt="$#,##0.00")
        wcell(ws2,r,5,row["Current $"],bg=bg,fmt="$#,##0.00")
        wcell(ws2,r,6,row["Cost Basis $"],bg=bg,fmt="$#,##0.00")
        wcell(ws2,r,7,row["Market Value $"],bg=bg,fmt="$#,##0.00")
        wcell(ws2,r,8,row["Unrealised $"],bg=bg,color=pc,fmt="$#,##0.00")
        wcell(ws2,r,9,row["Unrealised THB"],bg=bg,color=pc,fmt="#,##0")
        wcell(ws2,r,10,row["Net P&L %"],bg=bg,color=pc,fmt="0.00%")

    # ── Sheet 3: Transactions ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Transactions")
    ws3.sheet_view.showGridLines = False; ws3.sheet_properties.tabColor = "534AB7"
    for i, w in enumerate([6,12,8,10,10,10,12,12,12,12,28], 1): cw(ws3, i, w)
    titrow(ws3, 1, "TRANSACTION LEDGER", ncols=11)
    hdrrow(ws3, 2, ["ID","Date","Type","Ticker","Exchange","Shares","Price $","Gross $","Commission $","Total $","Note"])
    for i, tx in enumerate(cfg.get("transactions", [])):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws3, r, 19)
        tc = G if tx["type"] == "BUY" else R
        wcell(ws3,r,1,tx.get("id",""),bg=bg); wcell(ws3,r,2,tx.get("date",""),bg=bg)
        wcell(ws3,r,3,tx.get("type",""),bold=True,bg=bg,color=tc)
        wcell(ws3,r,4,tx.get("ticker",""),bold=True,bg=bg,color=HEADER)
        wcell(ws3,r,5,tx.get("exchange",""),bg=bg)
        wcell(ws3,r,6,tx.get("shares",0),bg=bg,fmt="#,##0")
        wcell(ws3,r,7,tx.get("price_usd",0),bg=bg,fmt="$#,##0.00")
        wcell(ws3,r,8,tx.get("gross_usd",0),bg=bg,fmt="$#,##0.00")
        wcell(ws3,r,9,tx.get("commission_usd",0),bg=bg,color=R,fmt="$#,##0.00")
        wcell(ws3,r,10,tx.get("total_usd",0),bold=True,bg=bg,fmt="$#,##0.00")
        wcell(ws3,r,11,tx.get("note",""),bg=bg)

    # ── Sheet 4: Dividend Tracker ────────────────────────────────────────────
    ws4 = wb.create_sheet("Dividend Tracker")
    ws4.sheet_view.showGridLines = False; ws4.sheet_properties.tabColor = G
    for i, w in enumerate([12,8,12,12,8,10,10,8,10,12,12,14], 1): cw(ws4, i, w)
    titrow(ws4, 1, "DIVIDEND TRACKER  --  History + Projected", ncols=12)
    hdrrow(ws4, 2, ["Period","Ticker","Ex-date","Pay-date","Shares","$/share",
                     "Gross $","WHT","Net $","Net THB est","KS App THB","Status"])
    sc_map = {"received": G, "pending": A, "projected": "3B8BD4", "MISSED": R}
    for i, d in enumerate(div_rows):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws4, r, 19)
        sc = sc_map.get(d.get("Status",""), DARK)
        fmts = [None,None,None,None,"#,##0","$#,##0.000","$#,##0.00",None,"$#,##0.00","#,##0",None,None]
        cols = ["Period","Ticker","Ex-date","Pay-date","Shares","$/share","Gross $","WHT","Net $","Net THB est","KS App THB","Status"]
        for j, (col, fmt) in enumerate(zip(cols, fmts)):
            xc = wcell(ws4, r, j+1, d.get(col,""), bg=bg, fmt=fmt,
                       color=(sc if col in ["Status","Ticker"] else DARK))
            if col == "Status" and d.get(col):
                xc.font = Font(bold=True, color="FFFFFF", size=10, name="Arial"); xc.fill = _fill(sc)

    # ── Sheet 5: WHT Reconciliation ──────────────────────────────────────────
    ws5 = wb.create_sheet("WHT Reconciliation")
    ws5.sheet_view.showGridLines = False; ws5.sheet_properties.tabColor = A
    for i, w in enumerate([10,8,8,12,12,12,12,12,14,14,20], 1): cw(ws5, i, w)
    titrow(ws5, 1, "WITHHOLDING TAX RECONCILIATION  --  30% vs 15% Treaty Rate", ncols=11)
    hdrrow(ws5, 2, ["Period","Ticker","Shares","Gross $","Net @30%","Net @15%",
                     "KS THB","KS USD est","Implied WHT","Verdict","Note"])
    vc_map = {"treaty_15": G, "default_30": R, "partial": A, "overpaid": R, "no_data": "888888"}
    for i, rec in enumerate(wht_records):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws5, r, 19)
        vc = vc_map.get(rec.verdict, A)
        for c, v, fmt in [
            (1,rec.period,None),(2,rec.ticker,None),(3,rec.shares,"#,##0"),
            (4,rec.gross_usd,"$#,##0.0000"),(5,rec.net_30pct,"$#,##0.0000"),(6,rec.net_15pct,"$#,##0.0000"),
            (7,rec.ks_thb or "missing",None),(8,f"${rec.ks_usd_est:.4f}" if rec.ks_usd_est else "n/a",None),
            (9,f"{rec.implied_wht*100:.1f}%" if rec.implied_wht else "n/a",None),
            (10,rec.verdict,None),(11,rec.note,None),
        ]:
            xc = wcell(ws5, r, c, v, bg=bg, fmt=fmt, color=(vc if c in [9,10] else DARK))
            if c == 10:
                xc.font = Font(bold=True, color="FFFFFF", size=10, name="Arial"); xc.fill = _fill(vc)

    # ── Sheet 6: Risk Metrics ────────────────────────────────────────────────
    ws6 = wb.create_sheet("Risk Metrics")
    ws6.sheet_view.showGridLines = False; ws6.sheet_properties.tabColor = "3B8BD4"
    nc6 = len(risk_tbl.index) + 1
    for i in range(1, nc6+1): cw(ws6, i, 20)
    titrow(ws6, 1, "INDIVIDUAL ASSET RISK METRICS  (Annualised)", ncols=nc6)
    hdrrow(ws6, 2, ["Metric"] + list(risk_tbl.index))
    fmt_map = {
        "Ann. Return":"0.00%","Ann. Volatility":"0.00%","Sharpe Ratio":"0.000",
        "Sortino Ratio":"0.000","Max Drawdown":"0.00%","Calmar Ratio":"0.000",
        "VaR 95%":"0.00%","CVaR 95%":"0.00%","Omega Ratio":"0.000","Semi-Volatility":"0.00%",
    }
    lower_better = {"Ann. Volatility","Max Drawdown","CVaR 95%","VaR 95%","Semi-Volatility"}
    for i, metric in enumerate(risk_tbl.columns):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws6, r, 19)
        wcell(ws6, r, 1, metric, bold=True, bg=bg)
        vals = risk_tbl[metric].dropna()
        best = vals.idxmin() if metric in lower_better else (vals.idxmax() if not vals.empty else None)
        for j, tkr in enumerate(risk_tbl.index):
            v = risk_tbl.loc[tkr, metric]; ib = tkr == best and pd.notna(v)
            wcell(ws6, r, 2+j, v, bg=f"{G}33" if ib else bg,
                  color=G if ib else DARK, fmt=fmt_map.get(metric,"0.000"))

    # ── Sheet 7: Optimal Weights ─────────────────────────────────────────────
    ws7 = wb.create_sheet("Optimal Weights")
    ws7.sheet_view.showGridLines = False; ws7.sheet_properties.tabColor = "7F77DD"
    nc7 = len(w_dict)+1; cw(ws7, 1, 22)
    for i in range(2, nc7+1): cw(ws7, i, 22)
    titrow(ws7, 1, "OPTIMAL PORTFOLIO WEIGHTS  --  All Strategies", ncols=nc7)
    hdrrow(ws7, 2, ["Asset"] + list(w_dict.keys()))
    tickers_all = list(list(w_dict.values())[0].keys() if isinstance(list(w_dict.values())[0], dict) else list(w_dict.values())[0].index) if w_dict else []
    for i, tkr in enumerate(tickers_all):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws7, r, 19)
        wcell(ws7, r, 1, tkr, bold=True, bg=bg)
        for j, wdf in enumerate(w_dict.values()):
            if isinstance(wdf, dict):
                # If it's a flat dictionary { 'AAPL': 0.5 }
                if tkr in wdf and not isinstance(wdf[tkr], dict):
                    v = float(wdf.get(tkr, 0.0))
                # If it's a nested dictionary { 'AAPL': {'weights': 0.5} }
                elif tkr in wdf and isinstance(wdf[tkr], dict):
                    v = float(wdf[tkr].get("weights", 0.0))
                # If it's oriented by column { 'weights': {'AAPL': 0.5} }
                elif "weights" in wdf and isinstance(wdf["weights"], dict):
                    v = float(wdf["weights"].get(tkr, 0.0))
                else:
                    v = 0.0
            else:
                # Original pandas logic
                v = float(wdf.loc[tkr, "weights"]) if hasattr(wdf, "index") and tkr in wdf.index else 0.0
            wcell(ws7, r, 2+j, v, bg=bg, color=G if v > 0.35 else DARK, fmt="0.0%")

    # ── Sheet 8: Riskfolio Charts ────────────────────────────────────────────
    ws8 = wb.create_sheet("Riskfolio Charts")
    ws8.sheet_view.showGridLines = False; ws8.sheet_properties.tabColor = HEADER
    titrow(ws8, 1, "RISKFOLIO-LIB VISUAL ANALYTICS")
    for key, anchor, pw, ph in [
        ("pie","A3",600,400), ("frontier","A26",700,460),
        ("risk_con","A54",640,380), ("hist","A77",680,380),
        ("drawdown","A100",700,480),
    ]:
        addimg(ws8, plots.get(key), anchor, pw, ph)

    # ── Sheet 9: Backtest ────────────────────────────────────────────────────
    ws9 = wb.create_sheet("Backtest")
    ws9.sheet_view.showGridLines = False; ws9.sheet_properties.tabColor = "534AB7"
    titrow(ws9, 1, "WALK-FORWARD BACKTEST RESULTS")
    if bt_metrics:
        hdrrow(ws9, 2, ["Strategy","Ann. Return","Ann. Vol","Sharpe","Max Drawdown","Calmar","Final $1"])
        for i, (s, m) in enumerate(bt_metrics.items()):
            r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws9, r, 19)
            wcell(ws9, r, 1, s.replace("_"," ").title(), bold=True, bg=bg)
            for j, (v, fmt) in enumerate([
                (m["ann_return"],"0.00%"), (m["ann_vol"],"0.00%"),
                (m["sharpe"],"0.000"), (m["max_drawdown"],"0.00%"),
                (m.get("calmar", np.nan),"0.000"), (m["final_equity"],"$0.000"),
            ], 1):
                safe = v if (not isinstance(v, float) or not np.isnan(v)) else "-"
                wcell(ws9, r, j+1, safe, bg=bg, fmt=fmt)
        addimg(ws9, plots.get("backtest_equity"), "A15", 700, 380)

    # ── Sheet 10: Generational Plan ──────────────────────────────────────────
    ws10 = wb.create_sheet("Generational Plan")
    ws10.sheet_view.showGridLines = False; ws10.sheet_properties.tabColor = G
    titrow(ws10, 1, "GENERATIONAL WEALTH PLAN  --  30-Year Monte Carlo")
    if milestones:
        hdrrow(ws10, 2, ["Year","p10 Value","p50 Value","p90 Value",
                          "p50 Real","p50 Income/mo","p10 Income/mo","P(>target)","Cum Div p50"])
        for i, (yr, ms) in enumerate(milestones.items()):
            r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws10, r, 19)
            wcell(ws10, r, 1, f"Year {yr}", bold=True, bg=bg)
            for j, (v, fmt) in enumerate([
                (ms["p10_value"],"$#,##0"), (ms["p50_value"],"$#,##0"),
                (ms["p90_value"],"$#,##0"), (ms["p50_real"],"$#,##0"),
                (ms["p50_income_m"],"$#,##0.00"), (ms["p10_income_m"],"$#,##0.00"),
                (ms["prob_above_target"],'0.0"%"'), (ms["p50_cum_div"],"$#,##0"),
            ], 1):
                wcell(ws10, r, j+2, v, bg=bg, fmt=fmt)
        addimg(ws10, plots.get("generational"), "A15", 700, 480)

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    try:
        from engine.macro_monitor import get_macro_data, get_macro_regime, get_risk_gauges
        _macro_data   = get_macro_data(cfg)
        _macro_regime = get_macro_regime(_macro_data)
        _macro_gauges = get_risk_gauges(_macro_data)
        _add_macro_sheet(wb, _macro_data, _macro_regime, _macro_gauges)
    except Exception as _me:
        import logging; logging.getLogger(__name__).warning("Macro sheet skipped: %s", _me)
    buf.seek(0)
    return buf.read()
