"""
riskfolio_autonomous.py  --  Portfolio Analytics Engine
=========================================================
Reads all portfolio data from config/portfolio.yaml.
To update your positions: edit portfolio.yaml, not this file.

SETUP
-----
    pip install riskfolio-lib yfinance openpyxl xlsxwriter pandas numpy pyyaml watchdog

RUN MANUALLY
------------
    python riskfolio_autonomous.py

AUTO-TRIGGER
------------
    python watchdog_runner.py   (watches config/portfolio.yaml, auto-runs on save)

SCHEDULE (cron -- Linux/macOS, every Monday 18:00)
----------------------------------------------------
    0 18 * * 1 python /path/to/riskfolio_autonomous.py >> /path/to/output/cron.log 2>&1

SCHEDULE (Windows Task Scheduler)
-----------------------------------
    Program  : python.exe
    Arguments: C:\\path\\to\\portfolio_app\\riskfolio_autonomous.py
    Trigger  : Weekly, Monday, 18:00
"""

import os
import sys
import warnings
import logging
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import yfinance as yf

try:
    import yaml
except ImportError:
    print("ERROR: pyyaml not installed. Run: pip install pyyaml")
    sys.exit(1)

try:
    import riskfolio as rp
except ImportError:
    print("ERROR: riskfolio-lib not installed. Run: pip install riskfolio-lib")
    sys.exit(1)

warnings.filterwarnings("ignore")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M",
)
log = logging.getLogger(__name__)

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).parent
CONFIG_FILE  = ROOT / "config" / "portfolio.yaml"
OUTPUT_DIR   = ROOT / "output"
PLOT_DIR     = OUTPUT_DIR / "plots"
OUTPUT_DIR.mkdir(exist_ok=True)
PLOT_DIR.mkdir(exist_ok=True)

TODAY        = datetime.today().strftime("%Y%m%d")
REPORT_PATH  = OUTPUT_DIR / f"portfolio_report_{TODAY}.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# 1. LOAD PORTFOLIO YAML
# ══════════════════════════════════════════════════════════════════════════════
def load_yaml() -> dict:
    if not CONFIG_FILE.exists():
        log.error("portfolio.yaml not found at: %s", CONFIG_FILE)
        sys.exit(1)
    with open(CONFIG_FILE, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    log.info("Loaded portfolio.yaml  (as_of: %s)", cfg["meta"].get("data_as_of", "unknown"))
    return cfg


def derive_holdings(cfg: dict) -> dict:
    """
    Compute current holdings (shares, avg cost, commissions) from transaction ledger.
    Returns dict keyed by ticker.
    """
    shares   = defaultdict(int)
    cost     = defaultdict(float)
    comm     = defaultdict(float)

    for tx in cfg.get("transactions", []):
        if tx["type"] == "BUY":
            ticker = tx["ticker"]
            shares[ticker] += tx["shares"]
            cost[ticker]   += tx["total_usd"]
            comm[ticker]   += tx["commission_usd"]

        elif tx["type"] == "SELL":
            ticker = tx["ticker"]
            shares[ticker] -= tx["shares"]
            # avg cost stays (for simplicity -- FIFO calc would be more precise)

    holdings = {}
    for ticker, sh in shares.items():
        if sh > 0:
            avg = cost[ticker] / sh
            holdings[ticker] = {
                "shares":        sh,
                "avg_cost":      round(avg, 4),
                "total_cost":    round(cost[ticker], 2),
                "total_comm":    round(comm[ticker], 2),
            }
    return holdings


def build_config(cfg: dict) -> dict:
    """Extract flat config values for use throughout the script."""
    s = cfg.get("settings", {})
    m = cfg.get("meta", {})
    return {
        "rf":           s.get("risk_free_rate_annual", 0.045),
        "wht":          s.get("wht_active", 0.30),
        "fx":           m.get("fx_usd_thb", 32.68),
        "data_start":   s.get("data_history_start", "2022-01-01"),
        "data_end":     datetime.today().strftime("%Y-%m-%d"),
        "points":       s.get("frontier_points", 50),
        "cash_usd":     cfg.get("cash", {}).get("usd", 0.0),
        "cash_thb":     cfg.get("cash", {}).get("thb", 0.0),
        "deposited":    cfg.get("cash", {}).get("total_deposited_usd", 0.0),
    }


# ══════════════════════════════════════════════════════════════════════════════
# 2. MARKET DATA
# ══════════════════════════════════════════════════════════════════════════════
def download_data(tickers: list, start: str, end: str):
    log.info("Downloading price data for %s ...", tickers)
    raw = yf.download(tickers, start=start, end=end,
                      auto_adjust=True, progress=False)
    if isinstance(raw.columns, pd.MultiIndex):
        prices = raw["Close"]
    else:
        prices = raw[["Close"]] if "Close" in raw.columns else raw
    prices  = prices.ffill().dropna(how="all")
    monthly = prices.resample("ME").last()
    returns = monthly.pct_change().dropna()
    returns = returns.dropna(axis=1, how="all")
    log.info("  %d months x %d tickers loaded.", len(returns), len(returns.columns))
    return monthly, returns


# ══════════════════════════════════════════════════════════════════════════════
# 3. PORTFOLIO OBJECT (Riskfolio-Lib v7)
# ══════════════════════════════════════════════════════════════════════════════
def build_portfolio(returns: pd.DataFrame) -> rp.Portfolio:
    port = rp.Portfolio(returns=returns)
    port.assets_stats(method_mu="hist", method_cov="ledoit")
    port.sht      = False
    port.upperlng = 1.0
    return port


# ══════════════════════════════════════════════════════════════════════════════
# 4. OPTIMISATIONS
# ══════════════════════════════════════════════════════════════════════════════
def run_optimisations(port: rp.Portfolio, rf_m: float) -> dict:
    strategies = [
        ("Max Sharpe (MV)",      "Classic", "MV",   "Sharpe"),
        ("Min Variance",         "Classic", "MV",   "MinRisk"),
        ("Max Sharpe (CVaR)",    "Classic", "CVaR", "Sharpe"),
        ("Min CVaR",             "Classic", "CVaR", "MinRisk"),
        ("Min Max Drawdown",     "Classic", "MDD",  "MinRisk"),
        ("Max Sharpe (Sortino)", "Classic", "SLPM", "Sharpe"),
    ]
    results = {}
    log.info("Running optimisations ...")
    for label, model, rm, obj in strategies:
        try:
            w = port.optimization(model=model, rm=rm, obj=obj,
                                  rf=rf_m, l=2, hist=True)
            if w is not None and not w.isnull().values.any():
                results[label] = w
                log.info("  [OK]   %s", label)
            else:
                log.warning("  [SKIP] %s", label)
        except Exception as exc:
            log.warning("  [FAIL] %s -- %s", label, exc)

    # Equal weight fallback
    n = len(port.returns.columns)
    results["Equal Weight"] = pd.DataFrame(
        [1/n]*n, index=port.returns.columns, columns=["weights"]
    )

    # HRP
    try:
        hport = rp.HCPortfolio(returns=port.returns)
        try:
            w_hrp = hport.optimization(
                model="HRP", codependence="pearson",
                rm="MV", rf=rf_m, linkage="ward", leaf_order=True,
            )
        except TypeError:
            w_hrp = hport.optimization(
                model="HRP", codependence="pearson",
                rm="MV", rf=rf_m, linkage="ward", max_k=10, leaf_order=True,
            )
        if w_hrp is not None:
            results["HRP (Risk Parity)"] = w_hrp
            log.info("  [OK]   HRP (Risk Parity)")
    except Exception as exc:
        log.warning("  [FAIL] HRP -- %s", exc)

    log.info("  %d strategies ready.", len(results))
    return results


# ══════════════════════════════════════════════════════════════════════════════
# 5. EFFICIENT FRONTIER
# ══════════════════════════════════════════════════════════════════════════════
def build_frontier(port: rp.Portfolio, rf_m: float) -> pd.DataFrame:
    log.info("Building efficient frontier ...")
    try:
        frontier = port.efficient_frontier(
            model="Classic", rm="MV", points=50, rf=rf_m, hist=True
        )
        log.info("  %d frontier portfolios.", frontier.shape[1])
        return frontier
    except Exception as exc:
        log.warning("  Frontier failed: %s", exc)
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# 6. RISK METRICS
# ══════════════════════════════════════════════════════════════════════════════
def compute_risk_table(returns: pd.DataFrame, rf: float) -> pd.DataFrame:
    rows = []
    for t in returns.columns:
        r = returns[t].dropna()
        if len(r) < 6:
            continue
        ann_ret = r.mean() * 12
        ann_vol = r.std() * np.sqrt(12)
        sr      = (r.mean()-rf) / r.std() * np.sqrt(12) if r.std() > 0 else np.nan
        neg     = r[r < 0]
        semi    = neg.std() * np.sqrt(12) if len(neg) > 1 else np.nan
        sortino = ((r.mean()-rf) / neg.std() * np.sqrt(12)
                   if len(neg) > 1 and neg.std() > 0 else np.nan)
        cum     = (1+r).cumprod()
        peak    = cum.cummax()
        dd      = (cum-peak)/peak
        mdd     = float(dd.min())
        calmar  = ann_ret / abs(mdd) if mdd != 0 else np.nan
        var95   = float(np.percentile(r, 5))
        mask    = r <= var95
        cvar95  = float(r[mask].mean()) if mask.any() else var95
        gains   = r[r > 0].sum()
        losses  = abs(r[r < 0].sum())
        omega   = gains/losses if losses > 0 else np.inf
        rows.append({
            "Ticker":          t,
            "Ann. Return":     ann_ret,
            "Ann. Volatility": ann_vol,
            "Sharpe Ratio":    sr,
            "Sortino Ratio":   sortino,
            "Max Drawdown":    mdd,
            "Calmar Ratio":    calmar,
            "VaR 95%":         var95,
            "CVaR 95%":        cvar95,
            "Omega Ratio":     omega,
            "Semi-Volatility": semi,
        })
    return pd.DataFrame(rows).set_index("Ticker")


# ══════════════════════════════════════════════════════════════════════════════
# 7. POSITION P&L  (uses YAML-derived holdings + live prices)
# ══════════════════════════════════════════════════════════════════════════════
def position_pnl(prices: pd.DataFrame, holdings: dict, fx: float) -> pd.DataFrame:
    rows = []
    for ticker, h in holdings.items():
        curr = float(prices[ticker].iloc[-1]) if ticker in prices.columns else 0.0
        sh, avg, total_cost = h["shares"], h["avg_cost"], h["total_cost"]
        mkt     = sh * curr
        unreal  = mkt - (sh * avg)
        net     = mkt - total_cost
        net_pct = net / total_cost if total_cost > 0 else 0.0
        rows.append({
            "Ticker":         ticker,
            "Shares":         sh,
            "Avg Cost $":     avg,
            "Current $":      round(curr, 2),
            "Cost Basis $":   round(total_cost, 2),
            "Market Value $": round(mkt, 2),
            "Unrealised $":   round(unreal, 2),
            "Unrealised THB": round(unreal * fx, 0),
            "Net P&L $":      round(net, 2),
            "Net P&L %":      round(net_pct, 4),
        })
    return pd.DataFrame(rows).set_index("Ticker")


# ══════════════════════════════════════════════════════════════════════════════
# 8. DIVIDEND PROJECTION  (from YAML instrument definitions)
# ══════════════════════════════════════════════════════════════════════════════
def build_dividend_projection(cfg: dict, holdings: dict, wht: float, fx: float) -> list:
    rows = []
    instruments = cfg.get("instruments", {})
    received    = cfg.get("dividends_received", [])

    # Historical received
    for d in received:
        ticker  = d.get("ticker", "")
        sh      = d.get("shares_eligible", 0)
        gross   = d.get("gross_usd_estimated", 0.0)
        net     = gross * (1 - wht)
        thb_ks  = d.get("thb_ks_app")
        rows.append({
            "Period":      d.get("period", ""),
            "Ticker":      ticker,
            "Ex-date":     d.get("ex_date", ""),
            "Pay-date":    d.get("pay_date", ""),
            "Shares":      sh,
            "$/share":     d.get("amount_per_share_usd", 0.0),
            "Gross $":     round(gross, 2),
            "WHT":         f"{wht*100:.0f}%",
            "Net $":       round(net, 2),
            "Net THB est": round(net * fx, 0),
            "KS App THB":  thb_ks if thb_ks else "n/a",
            "Status":      d.get("status", "received"),
        })

    # Upcoming from YAML definitions
    for ticker, inst in instruments.items():
        sh = holdings.get(ticker, {}).get("shares", 0)
        if sh == 0:
            continue
        for upcoming in inst.get("dividend_policy", {}).get("estimated_upcoming", []):
            eligible = upcoming.get("eligible_for_our_shares", True)
            amount   = upcoming.get("amount", upcoming.get("amount_per_share_usd", 0.0))
            gross    = sh * amount if eligible else 0.0
            net      = gross * (1 - wht)
            period   = upcoming.get("period", upcoming.get("ex", ""))
            rows.append({
                "Period":      period,
                "Ticker":      ticker,
                "Ex-date":     upcoming.get("ex", ""),
                "Pay-date":    upcoming.get("pay", ""),
                "Shares":      sh if eligible else 0,
                "$/share":     amount,
                "Gross $":     round(gross, 2),
                "WHT":         f"{wht*100:.0f}%",
                "Net $":       round(net, 2),
                "Net THB est": round(net * fx, 0),
                "KS App THB":  "n/a",
                "Status":      "projected" if eligible else "MISSED",
            })

    return rows


# ══════════════════════════════════════════════════════════════════════════════
# 9. RISKFOLIO PLOTS
# ══════════════════════════════════════════════════════════════════════════════
def _save(fig, name):
    p = str(PLOT_DIR / f"{name}.png")
    fig.savefig(p, dpi=140, bbox_inches="tight")
    plt.close(fig)
    return p


def generate_plots(port: rp.Portfolio, w_dict: dict, frontier: pd.DataFrame, rf_m: float) -> dict:
    plots = {}
    w_main = list(w_dict.values())[0]

    log.info("Generating Riskfolio-Lib plots ...")

    for name, fn in [
        ("pie",      lambda: _plot_pie(port, w_main)),
        ("frontier", lambda: _plot_frontier(port, w_main, frontier, rf_m)),
        ("risk_con", lambda: _plot_risk_con(port, w_main, rf_m)),
        ("hist",     lambda: _plot_hist(port, w_main)),
        ("drawdown", lambda: _plot_drawdown(port, w_main)),
        ("clusters", lambda: _plot_clusters(port)),
    ]:
        try:
            plots[name] = fn()
            log.info("  [OK] plot_%s", name)
        except Exception as e:
            log.warning("  plot_%s: %s", name, e)

    log.info("  %d / 6 plots saved to %s", len(plots), PLOT_DIR)
    return plots


def _plot_pie(port, w):
    fig, ax = plt.subplots(figsize=(8, 6))
    rp.plot_pie(w=w, title="Max Sharpe Allocation", others=0.05, nrow=25, cmap="tab20", ax=ax)
    return _save(fig, "pie")

def _plot_frontier(port, w, frontier, rf_m):
    if frontier.empty:
        raise ValueError("Empty frontier")
    fig, ax = plt.subplots(figsize=(10, 7))
    rp.plot_frontier(w_frontier=frontier, mu=port.mu, cov=port.cov,
                     returns=port.returns, rm="MV", rf=rf_m, alpha=0.05,
                     cmap="viridis", w=w, label="Max Sharpe", ax=ax)
    return _save(fig, "frontier")

def _plot_risk_con(port, w, rf_m):
    fig, ax = plt.subplots(figsize=(8, 5))
    rp.plot_risk_con(w=w, cov=port.cov, returns=port.returns,
                     rm="MV", rf=rf_m, alpha=0.05,
                     color="tab:blue", height=6, width=10, ax=ax)
    return _save(fig, "risk_con")

def _plot_hist(port, w):
    fig, ax = plt.subplots(figsize=(9, 5))
    rp.plot_hist(returns=port.returns, w=w, alpha=0.05, bins=50, height=6, width=10, ax=ax)
    return _save(fig, "hist")

def _plot_drawdown(port, w):
    fig, axes = plt.subplots(2, 1, figsize=(10, 8))
    rp.plot_drawdown(nav=None, w=w, returns=port.returns, alpha=0.05, height=8, width=10, ax=axes)
    return _save(fig, "drawdown")

def _plot_clusters(port):
    fig, ax = plt.subplots(figsize=(9, 5))
    rp.plot_clusters(returns=port.returns, codependence="pearson",
                     linkage="ward", k=None, max_k=10, leaf_order=True, ax=ax)
    return _save(fig, "clusters")


# ══════════════════════════════════════════════════════════════════════════════
# 10. EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════
def write_excel_report(
    prices, returns, risk_tbl, w_dict, pnl_df,
    frontier, plots, div_rows, cfg, holdings, conf
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    wb.remove(wb.active)

    # Style helpers
    def fill(h):  return PatternFill("solid", fgColor=h)
    def fnt(b=False, c="1A1A2E", s=10): return Font(bold=b, color=c, size=s, name="Arial")
    def aln(h="center"): return Alignment(horizontal=h, vertical="center")
    def bdr():
        s = Side(style="thin", color="CCCCDD")
        return Border(left=s, right=s, top=s, bottom=s)
    def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
    def rh(ws, row, h): ws.row_dimensions[row].height = h

    def wcell(ws, r, c, v, bold=False, color="1A1A2E", bg=None, fmt=None, size=10):
        x = ws.cell(r, c, v)
        x.font = fnt(bold, color, size); x.alignment = aln(); x.border = bdr()
        if bg:  x.fill          = fill(bg)
        if fmt: x.number_format = fmt
        return x

    def titrow(ws, row, text, ncols=9, bg="1A1A2E", h=32):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        x = ws[f"A{row}"]
        x.value = text; x.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
        x.fill = fill(bg); x.alignment = aln(); rh(ws, row, h)

    def hdrrow(ws, row, labels, bg="0F3460", h=22):
        rh(ws, row, h)
        for c, lbl in enumerate(labels, 1):
            x = ws.cell(row, c, lbl)
            x.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            x.fill = fill(bg); x.alignment = aln(); x.border = bdr()

    def addimg(ws, path, anchor, w=650, h=380):
        im = XLImage(path); im.width = w; im.height = h; ws.add_image(im, anchor)

    BG1 = "F2F4F8"; BG2 = "FFFFFF"
    G = "1D9E75"; A = "BA7517"; R = "A32D2D"

    # ── Sheet 1: Dashboard ───────────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = G
    for i, w in enumerate([10,16,14,14,14,14,14,14,14], 1): cw(ws, i, w)

    meta = cfg.get("meta", {})
    titrow(ws, 1,
           f"PORTFOLIO DASHBOARD  --  {meta.get('account_holder','')}  |  "
           f"Account {meta.get('account_id','')}  |  {datetime.today().strftime('%Y-%m-%d')}")

    # KS app snapshot
    snap = cfg.get("ks_app_snapshot_20260327", {})
    rh(ws, 3, 22)
    ws.merge_cells("A3:I3")
    c = ws.cell(3, 1, "KEY METRICS  (source: K CYBER TRADE app, 2026-03-27)")
    c.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    c.fill = fill("0F3460"); c.alignment = aln()

    metrics = [
        ("Market Value (THB)", f"฿{snap.get('market_value_thb',0):,.2f}"),
        ("Total Cost (THB)",   f"฿{snap.get('total_cost_thb',0):,.2f}"),
        ("Unrealised (THB)",   f"฿{snap.get('unrealized_thb',0):,.2f}"),
        ("Unrealised %",       f"{snap.get('unrealized_pct',0)*100:.2f}%"),
        ("Dividends (THB)",    f"฿{snap.get('total_dividends_thb',0):,.2f}"),
        ("Cash (USD)",         f"${conf['cash_usd']:.2f}"),
        ("Total Deposited $",  f"${conf['deposited']:,.2f}"),
        ("FX Rate",            f"{conf['fx']} THB/USD"),
    ]
    for i, (label, val) in enumerate(metrics):
        r = 4 + i; bg = BG1 if i%2==0 else BG2; rh(ws, r, 19)
        wcell(ws, r, 1, label, bold=True, bg=bg)
        for c2 in range(2, 10): ws.cell(r, c2).fill = fill(bg)
        ws.merge_cells(f"B{r}:I{r}")
        x = ws.cell(r, 2, val)
        x.font = fnt(bold=True, color=(R if "-" in str(val) else (G if "฿" in str(val) or "$" in str(val) else "1A1A2E")))
        x.fill = fill(bg); x.alignment = aln(); x.border = bdr()

    # ARCC missed dividend alert
    rh(ws, 13, 22)
    ws.merge_cells("A13:I13")
    x = ws.cell(13, 1, "ALERT: ARCC Q1 2026 dividend MISSED -- bought 2026-03-25, ex-date was 2026-03-12 (13 days early needed). "
                        "Next ARCC div: Q2 2026 est. ex 2026-06-12, pay 2026-06-30 -- 133 shares x $0.48 = $63.84 gross.")
    x.font = Font(bold=True, color="791F1F", size=10, name="Arial")
    x.fill = fill("FCEBEB"); x.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    x.border = bdr(); rh(ws, 13, 36)

    # ── Sheet 2: Position P&L ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Position P&L")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = G
    for i, w in enumerate([8,16,12,14,14,16,16,14,14,12], 1): cw(ws2, i, w)

    titrow(ws2, 1, f"POSITION P&L  --  Computed from YAML transaction ledger  --  {datetime.today().strftime('%Y-%m-%d')}")
    hdrrow(ws2, 2, ["#","Ticker","Shares","Avg Cost $","Current $",
                     "Cost Basis $","Market Value $","Unrealised $","Unrealised THB","Net P&L %"])

    for i, (tkr, row) in enumerate(pnl_df.iterrows()):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws2, r, 19)
        pnl_col = G if row["Net P&L $"] >= 0 else R
        wcell(ws2, r, 1, i+1, bg=bg)
        wcell(ws2, r, 2, tkr, bold=True, bg=bg, color="0F3460")
        wcell(ws2, r, 3, row["Shares"], bg=bg, fmt="#,##0")
        wcell(ws2, r, 4, row["Avg Cost $"], bg=bg, fmt="$#,##0.00")
        wcell(ws2, r, 5, row["Current $"], bg=bg, fmt="$#,##0.00")
        wcell(ws2, r, 6, row["Cost Basis $"], bg=bg, fmt="$#,##0.00")
        wcell(ws2, r, 7, row["Market Value $"], bg=bg, fmt="$#,##0.00")
        wcell(ws2, r, 8, row["Unrealised $"], bg=bg, color=pnl_col, fmt="$#,##0.00")
        wcell(ws2, r, 9, row["Unrealised THB"], bg=bg, color=pnl_col, fmt="#,##0")
        wcell(ws2, r, 10, row["Net P&L %"], bg=bg, color=pnl_col, fmt="0.00%")

    # Totals row
    r_tot = 3 + len(pnl_df)
    rh(ws2, r_tot, 20)
    tot_cost = pnl_df["Cost Basis $"].sum()
    tot_mkt  = pnl_df["Market Value $"].sum()
    tot_unr  = pnl_df["Unrealised $"].sum()
    tot_thb  = pnl_df["Unrealised THB"].sum()
    tot_pct  = tot_unr / tot_cost if tot_cost > 0 else 0
    for c2 in range(1, 11):
        ws2.cell(r_tot, c2).fill = fill("0F3460")
    wcell(ws2, r_tot, 2, "TOTAL", bold=True, color="FFFFFF", bg="0F3460")
    wcell(ws2, r_tot, 6, tot_cost, bold=True, color="FFFFFF", bg="0F3460", fmt="$#,##0.00")
    wcell(ws2, r_tot, 7, tot_mkt,  bold=True, color="FFFFFF", bg="0F3460", fmt="$#,##0.00")
    wcell(ws2, r_tot, 8, tot_unr,  bold=True, color=(G if tot_unr>=0 else R), bg="0F3460", fmt="$#,##0.00")
    wcell(ws2, r_tot, 9, tot_thb,  bold=True, color=(G if tot_thb>=0 else R), bg="0F3460", fmt="#,##0")
    wcell(ws2, r_tot, 10, tot_pct, bold=True, color=(G if tot_pct>=0 else R), bg="0F3460", fmt="0.00%")

    # ── Sheet 3: Transaction Ledger ──────────────────────────────────────────
    ws3 = wb.create_sheet("Transactions")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "534AB7"
    for i, w in enumerate([6,12,8,10,12,10,10,12,12,10,28], 1): cw(ws3, i, w)

    titrow(ws3, 1, "TRANSACTION LEDGER  --  Source: portfolio.yaml", ncols=11)
    hdrrow(ws3, 2, ["ID","Date","Type","Ticker","Exchange","Shares",
                     "Price $","Gross $","Commission $","Total $","Note"])

    for i, tx in enumerate(cfg.get("transactions", [])):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws3, r, 19)
        t_col = G if tx["type"] == "BUY" else R
        wcell(ws3, r, 1, tx.get("id",""), bg=bg)
        wcell(ws3, r, 2, tx.get("date",""), bg=bg)
        wcell(ws3, r, 3, tx.get("type",""), bold=True, bg=bg, color=t_col)
        wcell(ws3, r, 4, tx.get("ticker",""), bold=True, bg=bg, color="0F3460")
        wcell(ws3, r, 5, tx.get("exchange",""), bg=bg)
        wcell(ws3, r, 6, tx.get("shares",0), bg=bg, fmt="#,##0")
        wcell(ws3, r, 7, tx.get("price_usd",0), bg=bg, fmt="$#,##0.00")
        wcell(ws3, r, 8, tx.get("gross_usd",0), bg=bg, fmt="$#,##0.00")
        wcell(ws3, r, 9, tx.get("commission_usd",0), bg=bg, color=R, fmt="$#,##0.00")
        wcell(ws3, r, 10, tx.get("total_usd",0), bold=True, bg=bg, fmt="$#,##0.00")
        wcell(ws3, r, 11, tx.get("note",""), bg=bg)

    # ── Sheet 4: Dividend Tracker ────────────────────────────────────────────
    ws4 = wb.create_sheet("Dividend Tracker")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = G
    for i, w in enumerate([12,8,12,12,8,10,10,8,10,12,12,14], 1): cw(ws4, i, w)

    titrow(ws4, 1, "DIVIDEND TRACKER  --  History + Projected  (WHT applied to projections)", ncols=12)
    hdrrow(ws4, 2, ["Period","Ticker","Ex-date","Pay-date","Shares",
                     "$/share","Gross $","WHT","Net $","Net THB est",
                     "KS App THB","Status"])

    status_colors = {"received": G, "pending": A, "projected": "3B8BD4", "MISSED": R}
    for i, d in enumerate(div_rows):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws4, r, 19)
        sc = status_colors.get(d.get("Status",""), "1A1A2E")
        cols = ["Period","Ticker","Ex-date","Pay-date","Shares",
                "$/share","Gross $","WHT","Net $","Net THB est","KS App THB","Status"]
        fmts = [None, None, None, None, "#,##0", "$#,##0.000",
                "$#,##0.00", None, "$#,##0.00", "#,##0", None, None]
        for j, (col, fmt) in enumerate(zip(cols, fmts)):
            c_col = (sc if col in ["Status","Ticker"] else "1A1A2E")
            v = d.get(col, "")
            xcell = wcell(ws4, r, j+1, v, bg=bg, color=c_col, fmt=fmt)
            if col == "Status" and v:
                xcell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
                xcell.fill = fill(sc[:6] if len(sc)>6 else sc)

    # ARCC Q2 upcoming callout
    r_note = 3 + len(div_rows) + 2
    ws4.merge_cells(f"A{r_note}:L{r_note}")
    x = ws4.cell(r_note, 1,
                 "NEXT KEY DATE: ARCC Q2 2026 dividend -- estimated ex-date 2026-06-12, "
                 "payment 2026-06-30 -- 133 shares x $0.48 = $63.84 gross / "
                 f"${63.84*(1-conf['wht']):.2f} net after {conf['wht']*100:.0f}% WHT = "
                 f"~฿{63.84*(1-conf['wht'])*conf['fx']:,.0f} THB")
    x.font = Font(bold=False, color="085041", size=10, name="Arial")
    x.fill = fill("E1F5EE"); x.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    x.border = bdr(); rh(ws4, r_note, 36)

    # ── Sheet 5: Risk Metrics ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Risk Metrics")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = "3B8BD4"
    nc5 = len(risk_tbl.index) + 1
    for i in range(1, nc5+1): cw(ws5, i, 20)
    titrow(ws5, 1, "INDIVIDUAL ASSET RISK METRICS  (Annualised, from monthly total returns)", ncols=nc5)
    hdrrow(ws5, 2, ["Metric"] + list(risk_tbl.index))

    fmt_map = {
        "Ann. Return": "0.00%", "Ann. Volatility": "0.00%",
        "Sharpe Ratio": "0.000", "Sortino Ratio": "0.000",
        "Max Drawdown": "0.00%", "Calmar Ratio": "0.000",
        "VaR 95%": "0.00%", "CVaR 95%": "0.00%",
        "Omega Ratio": "0.000", "Semi-Volatility": "0.00%",
    }
    lower_better = {"Ann. Volatility","Max Drawdown","CVaR 95%","VaR 95%","Semi-Volatility"}
    for i, metric in enumerate(risk_tbl.columns):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws5, r, 19)
        wcell(ws5, r, 1, metric, bold=True, bg=bg)
        vals = risk_tbl[metric].dropna()
        best = vals.idxmin() if metric in lower_better else (vals.idxmax() if not vals.empty else None)
        for j, tkr in enumerate(risk_tbl.index):
            v = risk_tbl.loc[tkr, metric]
            ib = tkr == best and pd.notna(v)
            wcell(ws5, r, 2+j, v, bg=f"{G}33" if ib else bg,
                  color=G if ib else "1A1A2E", fmt=fmt_map.get(metric,"0.000"))

    # ── Sheet 6: Optimal Weights ─────────────────────────────────────────────
    ws6 = wb.create_sheet("Optimal Weights")
    ws6.sheet_view.showGridLines = False
    ws6.sheet_properties.tabColor = "7F77DD"
    nc6 = len(w_dict)+1
    cw(ws6, 1, 22)
    for i in range(2, nc6+1): cw(ws6, i, 20)
    titrow(ws6, 1, "OPTIMAL PORTFOLIO WEIGHTS -- All Strategies", ncols=nc6)
    hdrrow(ws6, 2, ["Asset"] + list(w_dict.keys()))

    for i, tkr in enumerate(returns.columns):
        r = 3+i; bg = BG1 if i%2==0 else BG2; rh(ws6, r, 19)
        wcell(ws6, r, 1, tkr, bold=True, bg=bg)
        for j, wdf in enumerate(w_dict.values()):
            v = float(wdf.loc[tkr,"weights"]) if tkr in wdf.index else 0.0
            wcell(ws6, r, 2+j, v, bg=bg, color=G if v>0.35 else "1A1A2E", fmt="0.0%")

    # ── Sheet 7: Charts ──────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Riskfolio Charts")
    ws7.sheet_view.showGridLines = False
    ws7.sheet_properties.tabColor = "0F3460"
    titrow(ws7, 1, "RISKFOLIO-LIB VISUAL ANALYTICS  --  Pie / Frontier / Risk / Hist / Drawdown / Clusters")
    for key, anchor, pw, ph in [
        ("pie","A3",600,400), ("frontier","A26",700,460),
        ("risk_con","A54",640,380), ("hist","A77",680,380),
        ("drawdown","A100",700,480), ("clusters","A128",680,380),
    ]:
        if key in plots: addimg(ws7, plots[key], anchor, pw, ph)

    # ── Sheet 8: Efficient Frontier ──────────────────────────────────────────
    if not frontier.empty:
        ws8 = wb.create_sheet("Frontier Data")
        ws8.sheet_view.showGridLines = False
        ws8.sheet_properties.tabColor = "0F3460"
        for i, w in enumerate([12,18,18,16], 1): cw(ws8, i, w)
        titrow(ws8, 1, "MEAN-VARIANCE EFFICIENT FRONTIER", ncols=4)
        hdrrow(ws8, 2, ["Portfolio #","Volatility (Ann.)","Return (Ann.)","Sharpe Ratio"])
        ef_r   = returns.values @ frontier.values
        ef_ret = ef_r.mean(axis=0)*12; ef_vol = ef_r.std(axis=0)*np.sqrt(12)
        ef_sh  = (ef_ret - conf["rf"]) / np.where(ef_vol>0, ef_vol, np.nan)
        best   = int(np.nanargmax(ef_sh))
        for i in range(len(ef_ret)):
            r = 3+i; bg = f"{G}22" if i==best else (BG1 if i%2==0 else BG2); rh(ws8, r, 17)
            for c2, v, fmt in [(1,i+1,None),(2,ef_vol[i],"0.00%"),(3,ef_ret[i],"0.00%"),(4,ef_sh[i],"0.000")]:
                wcell(ws8, r, c2, v, bg=bg, fmt=fmt)

    wb.save(str(REPORT_PATH))
    log.info("Report saved --> %s", REPORT_PATH)
    return str(REPORT_PATH)


# ══════════════════════════════════════════════════════════════════════════════
# 11. MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    log.info("=" * 62)
    log.info("  PORTFOLIO ANALYTICS ENGINE  --  STARTING")
    log.info("  riskfolio: %s  |  config: %s", rp.__version__, CONFIG_FILE.name)
    log.info("=" * 62)

    # Load YAML
    cfg      = load_yaml()
    conf     = build_config(cfg)
    holdings = derive_holdings(cfg)
    rf_m     = conf["rf"] / 12

    log.info("Holdings derived from %d transactions:", len(cfg.get("transactions", [])))
    for tkr, h in holdings.items():
        log.info("  %s: %d shares @ avg $%.2f  (total cost $%.2f)",
                 tkr, h["shares"], h["avg_cost"], h["total_cost"])

    tickers = list(holdings.keys())
    if not tickers:
        log.error("No holdings found in portfolio.yaml transactions.")
        sys.exit(1)

    # Market data
    prices, returns = download_data(tickers, conf["data_start"], conf["data_end"])

    # Portfolio object
    port     = build_portfolio(returns)
    w_dict   = run_optimisations(port, rf_m)
    frontier = build_frontier(port, rf_m)
    risk_tbl = compute_risk_table(returns, rf_m)
    pnl_df   = position_pnl(prices, holdings, conf["fx"])
    div_rows = build_dividend_projection(cfg, holdings, conf["wht"], conf["fx"])
    plots    = generate_plots(port, w_dict, frontier, rf_m)

    path = write_excel_report(
        prices, returns, risk_tbl, w_dict, pnl_df,
        frontier, plots, div_rows, cfg, holdings, conf
    )

    log.info("=" * 62)
    log.info("  DONE  -->  %s", path)
    log.info("=" * 62)

    # Console summary
    total_mkt   = pnl_df["Market Value $"].sum()
    total_unr   = pnl_df["Unrealised $"].sum()
    total_unr_t = pnl_df["Unrealised THB"].sum()
    div_thb     = cfg.get("ks_app_snapshot_20260327", {}).get("total_dividends_thb", 0)

    log.info("  Portfolio value  : $%,.2f  (฿%,.0f)", total_mkt, total_mkt * conf["fx"])
    log.info("  Unrealised P&L   : $%,.2f  (฿%,.0f)", total_unr, total_unr_t)
    log.info("  Dividends (KS)   : ฿%,.2f", div_thb)
    log.info("  Cash (USD)       : $%.2f", conf["cash_usd"])
    log.info("  ARCC Q2 2026 est : $%.2f gross  $%.2f net  ฿%,.0f  (ex ~2026-06-12)",
             63.84, 63.84*(1-conf["wht"]), 63.84*(1-conf["wht"])*conf["fx"])


if __name__ == "__main__":
    main()
